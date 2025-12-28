#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║     AQUABRAIN TENDER GENERATOR - Professional Hebrew BoQ                     ║
║     Data Extraction → Contract Documentation                                 ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import json
import os
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, Border, Side, PatternFill, Protection
    )
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run(['pip3', 'install', 'pandas', 'openpyxl', '-q'])
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, Border, Side, PatternFill, Protection
    )
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

INPUT_JSON = '/mnt/c/Temp/AquaBrain/boq_final.json'
OUTPUT_XLSX = '/mnt/c/Users/נימרודעופר/Desktop/1783-nim - Standard/Tender_BoQ_Professional.xlsx'

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
CHAPTER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
FOOTER_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADER_FONT = Font(name='Arial', size=12, bold=True, color="FFFFFF")
CHAPTER_FONT = Font(name='Arial', size=11, bold=True, color="FFFFFF")
ITEM_FONT = Font(name='Arial', size=10)
FOOTER_FONT = Font(name='Arial', size=9, bold=True, color="C00000")

# ═══════════════════════════════════════════════════════════════════════════════
# LOAD DATA
# ═══════════════════════════════════════════════════════════════════════════════

def load_extraction_data():
    """Load the extracted BoQ JSON."""
    with open(INPUT_JSON, 'r', encoding='latin-1') as f:
        content = f.read()
    content = content.encode('utf-8', errors='replace').decode('utf-8')
    return json.loads(content)

# ═══════════════════════════════════════════════════════════════════════════════
# BOQ STRUCTURE
# ═══════════════════════════════════════════════════════════════════════════════

def build_boq_data(data):
    """Build the BoQ structure with Hebrew specifications."""

    # Extract quantities from JSON
    spk_types = data.get('sprinkler_types', {})
    valve_types = data.get('valve_types', {})
    pipes = data.get('pipes', {})

    # Calculate totals
    total_sprinklers = data.get('sprinkler_count', 0)
    total_valves = data.get('valve_count', 0)

    # Sprinkler breakdown
    pendent_count = sum(v for k, v in spk_types.items() if 'PEND' in k.upper())
    sidewall_count = sum(v for k, v in spk_types.items() if 'SIDEWALL' in k.upper())
    concealed_count = sum(v for k, v in spk_types.items() if 'CONCEALED' in k.upper())
    other_spk = total_sprinklers - pendent_count - sidewall_count - concealed_count

    # Valve breakdown
    chk_valves = sum(v for k, v in valve_types.items() if 'CHK' in k.upper() or 'CHECK' in k.upper())
    f_valves = sum(v for k, v in valve_types.items() if 'F-VALVE' in k.upper() or 'F_VALVE' in k.upper())
    zone_valves = sum(v for k, v in valve_types.items() if 'ZONE' in k.upper() or 'CONTROL' in k.upper())

    # Estimate zone controls (if not found, estimate based on typical 1 per floor/zone)
    zone_control_count = zone_valves if zone_valves > 0 else max(1, total_valves // 20)
    fire_hose_count = max(1, total_valves // 50)  # Estimate

    # Pipe lengths (from extraction or estimate)
    main_pipe_len = sum(float(v) for k, v in pipes.items() if int(k) == 4) / 1000 if pipes else 0
    branch_pipe_len = sum(float(v) for k, v in pipes.items() if int(k) in [2, 5, 6]) / 1000 if pipes else 0

    # If no pipe data, estimate from sprinkler count
    if main_pipe_len == 0:
        main_pipe_len = total_sprinklers * 3.5  # ~3.5m main per sprinkler
    if branch_pipe_len == 0:
        branch_pipe_len = total_sprinklers * 2.0  # ~2m branch per sprinkler

    # Build BoQ rows
    boq_items = [
        # Chapter 0: General
        {'chapter': '0', 'item': '', 'description': 'פרק 0: כללי ומסירה', 'unit': '', 'qty': '', 'price': '', 'total': '', 'is_chapter': True},
        {'chapter': '0', 'item': '0.1', 'description': 'תיעוד מסירה (AS MADE) - המצאת תעודה המאשרת תקינות + תוכניות עדות', 'unit': 'קומפלט', 'qty': 1, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '0', 'item': '0.2', 'description': 'בדיקות לחץ - בדיקת המערכת בלחץ אוויר (10 אטמ) ומים (15 אטמ) לפי ת"י 1596', 'unit': 'קומפלט', 'qty': 1, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '0', 'item': '0.3', 'description': 'אישור מכבי אש - ליווי הליך הרישוי מול הרשות המקומית', 'unit': 'קומפלט', 'qty': 1, 'price': '', 'total': '', 'is_chapter': False},

        # Chapter 1: Piping
        {'chapter': '1', 'item': '', 'description': 'פרק 1: צנרת ואביזרים', 'unit': '', 'qty': '', 'price': '', 'total': '', 'is_chapter': True},
        {'chapter': '1', 'item': '1.1', 'description': 'צינורות פלדה מגולוונים סקדיול 40 (ללא תפר, הברגה) - קטרים 1"-2"', 'unit': 'מ"א', 'qty': round(branch_pipe_len, 1), 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '1', 'item': '1.2', 'description': 'צינורות פלדה מגולוונים סקדיול 10 (חיבורי GROOVED/QUICK-UP) - קטרים 2.5"-4"', 'unit': 'מ"א', 'qty': round(main_pipe_len, 1), 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '1', 'item': '1.3', 'description': 'צנרת גמישה (שרוול נירוסטה שזור 304) - עד 1.5 מטר, מאושר UL/FM', 'unit': 'יח\'', 'qty': total_sprinklers, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '1', 'item': '1.4', 'description': 'אביזרי צנרת (ברזים, רדוקציות, תפסנים, ברגים) - כלול במחיר הצנרת', 'unit': 'קומפלט', 'qty': 1, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '1', 'item': '1.5', 'description': 'צביעת צנרת ואביזרים (אדום RAL 3000 לסמוי, לבן RAL 9010 לגלוי)', 'unit': 'קומפלט', 'qty': 1, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '1', 'item': '1.6', 'description': 'תליות ותמיכות (כולל מעורר רעידות SEISMIC כנדרש)', 'unit': 'קומפלט', 'qty': 1, 'price': '', 'total': '', 'is_chapter': False},

        # Chapter 2: Valves
        {'chapter': '2', 'item': '', 'description': 'פרק 2: מגופים ואביזרי שליטה', 'unit': '', 'qty': '', 'price': '', 'total': '', 'is_chapter': True},
        {'chapter': '2', 'item': '2.1', 'description': 'סט הפעלה דירתי/אזורי (ZONE CONTROL) - כולל ברז ראשי, מד זרימה, ברז בדיקה וברז ניקוז', 'unit': 'יח\'', 'qty': zone_control_count, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '2', 'item': '2.2', 'description': 'שסתום חד כיווני (CHECK VALVE) - ברונזה מאושר UL/FM', 'unit': 'יח\'', 'qty': chk_valves, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '2', 'item': '2.3', 'description': 'ברז כדורי (BALL VALVE) - ידית T, ברונזה מאושר', 'unit': 'יח\'', 'qty': f_valves, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '2', 'item': '2.4', 'description': 'עמדת כיבוי אש (גלגלון 25 מ\') - קומפלט כולל ארון, צינור, מזנק ומטף 6 ק"ג', 'unit': 'יח\'', 'qty': fire_hose_count, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '2', 'item': '2.5', 'description': 'שסתום רטוב (WET ALARM VALVE) - כולל מעורר לחץ ופעמון מים', 'unit': 'יח\'', 'qty': max(1, zone_control_count // 4), 'price': '', 'total': '', 'is_chapter': False},

        # Chapter 3: Sprinklers
        {'chapter': '3', 'item': '', 'description': 'פרק 3: מתזים (ספרינקלרים)', 'unit': '', 'qty': '', 'price': '', 'total': '', 'is_chapter': True},
        {'chapter': '3', 'item': '3.1', 'description': 'ספרינקלר תלוי (PENDENT) - תגובה מהירה QR, K-Factor=5.6, 68°C, ציפוי כרום', 'unit': 'יח\'', 'qty': pendent_count, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '3', 'item': '3.2', 'description': 'ספרינקלר צידי (SIDEWALL) - תגובה מהירה QR, K-Factor=5.6, 68°C, ציפוי לבן', 'unit': 'יח\'', 'qty': sidewall_count, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '3', 'item': '3.3', 'description': 'ספרינקלר סמוי (CONCEALED) - כולל מכסה דקורטיבי לבן, 57°C', 'unit': 'יח\'', 'qty': concealed_count if concealed_count > 0 else '-', 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '3', 'item': '3.4', 'description': 'ספרינקלר זקוף (UPRIGHT) - לשימוש בחדרי מכונות, 93°C', 'unit': 'יח\'', 'qty': other_spk if other_spk > 0 else '-', 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '3', 'item': '3.5', 'description': 'רוזטות דקורטיביות (ESCUTCHEON) - לבן/כרום, מאושר', 'unit': 'יח\'', 'qty': total_sprinklers, 'price': '', 'total': '', 'is_chapter': False},

        # Chapter 4: Accessories
        {'chapter': '4', 'item': '', 'description': 'פרק 4: אביזרים ושונות', 'unit': '', 'qty': '', 'price': '', 'total': '', 'is_chapter': True},
        {'chapter': '4', 'item': '4.1', 'description': 'שילוט וסימון (תוויות זיהוי, חצים כיוון זרימה, שלטי אזהרה)', 'unit': 'קומפלט', 'qty': 1, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '4', 'item': '4.2', 'description': 'פתחים וקידוחים בבטון/בלוקים (כולל איטום מעברי אש EI)', 'unit': 'קומפלט', 'qty': 1, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '4', 'item': '4.3', 'description': 'מפתח כיבוי + תיק כלים (לשימוש מכבי אש)', 'unit': 'סט', 'qty': 1, 'price': '', 'total': '', 'is_chapter': False},
        {'chapter': '4', 'item': '4.4', 'description': 'חומרי מילוי ואיטום (טפלון, דבק אנאארובי, רצועות גומי)', 'unit': 'קומפלט', 'qty': 1, 'price': '', 'total': '', 'is_chapter': False},
    ]

    return boq_items

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def generate_excel(boq_items, data):
    """Generate professional Hebrew Excel BoQ."""

    wb = Workbook()
    ws = wb.active
    ws.title = "כתב כמויות"

    # Set RTL direction
    ws.sheet_view.rightToLeft = True

    # Column widths
    ws.column_dimensions['A'].width = 8   # פרק
    ws.column_dimensions['B'].width = 8   # סעיף
    ws.column_dimensions['C'].width = 65  # תיאור
    ws.column_dimensions['D'].width = 10  # יחידה
    ws.column_dimensions['E'].width = 12  # כמות
    ws.column_dimensions['F'].width = 14  # מחיר יחידה
    ws.column_dimensions['G'].width = 16  # סה"כ

    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "כתב כמויות - מערכת כיבוי אש (ספרינקלרים)"
    title_cell.font = Font(name='Arial', size=16, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Project info
    ws.merge_cells('A2:G2')
    info_cell = ws['A2']
    info_cell.value = f"פרויקט: 1783-nim | תאריך: {datetime.now().strftime('%d/%m/%Y')} | מקור: AutoCAD 2026 Extraction"
    info_cell.font = Font(name='Arial', size=10, italic=True)
    info_cell.alignment = Alignment(horizontal='center')

    # Headers (Row 4)
    headers = ['פרק', 'סעיף', 'תיאור הפריט / מפרט טכני', 'יחידה', 'כמות', 'מחיר יחידה (₪)', 'סה"כ (₪)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 25

    # Data rows
    row_num = 5
    for item in boq_items:
        if item['is_chapter']:
            # Chapter row
            ws.merge_cells(f'A{row_num}:G{row_num}')
            cell = ws.cell(row=row_num, column=1, value=item['description'])
            cell.font = CHAPTER_FONT
            cell.fill = CHAPTER_FILL
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = THIN_BORDER
            ws.row_dimensions[row_num].height = 22
        else:
            # Item row
            ws.cell(row=row_num, column=1, value=item['chapter']).border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=item['item']).border = THIN_BORDER

            desc_cell = ws.cell(row=row_num, column=3, value=item['description'])
            desc_cell.font = ITEM_FONT
            desc_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            desc_cell.border = THIN_BORDER

            ws.cell(row=row_num, column=4, value=item['unit']).border = THIN_BORDER
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')

            qty_cell = ws.cell(row=row_num, column=5, value=item['qty'])
            qty_cell.border = THIN_BORDER
            qty_cell.alignment = Alignment(horizontal='center')

            price_cell = ws.cell(row=row_num, column=6, value=item['price'])
            price_cell.border = THIN_BORDER
            price_cell.alignment = Alignment(horizontal='center')

            total_cell = ws.cell(row=row_num, column=7, value=item['total'])
            total_cell.border = THIN_BORDER
            total_cell.alignment = Alignment(horizontal='center')

            # Alternate row coloring
            if row_num % 2 == 0:
                for col in range(1, 8):
                    if not item['is_chapter']:
                        ws.cell(row=row_num, column=col).fill = ALT_ROW_FILL

            ws.row_dimensions[row_num].height = 35

        row_num += 1

    # Summary row
    row_num += 1
    ws.merge_cells(f'A{row_num}:F{row_num}')
    summary_cell = ws.cell(row=row_num, column=1, value='סה"כ לפני מע"מ:')
    summary_cell.font = Font(name='Arial', size=12, bold=True)
    summary_cell.alignment = Alignment(horizontal='left')
    ws.cell(row=row_num, column=7).border = THIN_BORDER
    ws.cell(row=row_num, column=7).font = Font(name='Arial', size=12, bold=True)

    row_num += 1
    ws.merge_cells(f'A{row_num}:F{row_num}')
    vat_cell = ws.cell(row=row_num, column=1, value='מע"מ (17%):')
    vat_cell.font = Font(name='Arial', size=11)

    row_num += 1
    ws.merge_cells(f'A{row_num}:F{row_num}')
    grand_cell = ws.cell(row=row_num, column=1, value='סה"כ כולל מע"מ:')
    grand_cell.font = Font(name='Arial', size=14, bold=True, color="C00000")
    ws.cell(row=row_num, column=7).font = Font(name='Arial', size=14, bold=True, color="C00000")

    # Footer note
    row_num += 2
    ws.merge_cells(f'A{row_num}:G{row_num}')
    footer_cell = ws.cell(row=row_num, column=1)
    footer_cell.value = (
        "למען הסר ספק: המחירים כוללים ביצוע פתחים, קידוחים בבטון, איטום מעברים EI, "
        "תליות עזר וקונסטרוקציה, שילוט וסימון, בדיקות וניסויים - הכל מושלם (קומפלט) לפי תקן ישראלי 1596 ו-NFPA 13."
    )
    footer_cell.font = FOOTER_FONT
    footer_cell.fill = FOOTER_FILL
    footer_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws.row_dimensions[row_num].height = 40

    # Statistics note
    row_num += 2
    ws.merge_cells(f'A{row_num}:G{row_num}')
    stats_cell = ws.cell(row=row_num, column=1)
    spk_count = data.get('sprinkler_count', 0)
    valve_count = data.get('valve_count', 0)
    stats_cell.value = f"נתוני חילוץ: {spk_count} ספרינקלרים | {valve_count} מגופים | מקור: AquaBrain AutoCAD Extractor v2.1"
    stats_cell.font = Font(name='Arial', size=9, italic=True, color="666666")
    stats_cell.alignment = Alignment(horizontal='center')

    return wb

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║     AQUABRAIN TENDER GENERATOR - Hebrew Professional BoQ    ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()

    # Load data
    print("[1/4] Loading extraction data...")
    data = load_extraction_data()
    print(f"      ✓ Found {data.get('sprinkler_count', 0)} sprinklers, {data.get('valve_count', 0)} valves")

    # Build BoQ
    print("[2/4] Building BoQ structure...")
    boq_items = build_boq_data(data)
    print(f"      ✓ Created {len(boq_items)} line items")

    # Generate Excel
    print("[3/4] Generating Excel workbook...")
    wb = generate_excel(boq_items, data)
    print("      ✓ Workbook created with formatting")

    # Save
    print("[4/4] Saving to Desktop...")
    os.makedirs(os.path.dirname(OUTPUT_XLSX), exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"      ✓ Saved to: {OUTPUT_XLSX}")

    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║     ✓ TENDER DOCUMENT GENERATED SUCCESSFULLY                ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()
    print(f"Output: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
